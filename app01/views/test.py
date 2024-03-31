import json
import os

from django.shortcuts import render,redirect,HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from openpyxl import load_workbook

from app01 import models

def test_list(request):
    if request.method == "GET":
        return render(request,"test_list.html")
    print(request)
    print(request.POST)
    print(request.FILES)
    file_object = request.FILES.get("avatar")
    print(file_object.name)
    print(type(file_object))
    
    # excel文件
    # wb = load_workbook(file_object)
    # sheet = wb.worksheets[0]
    # cell = sheet.cell(1,1)
    # print(cell.value)
    # for row in sheet.iter_rows(min_row=2):
    #     text = row[0].value
    #     print(text)

    from django.core.files.uploadedfile import InMemoryUploadedFile
    file_path = os.path.join("media",file_object.name)
    with open(file_path,mode="wb") as f:
        for chunk in file_object.chunks():
            f.write(chunk)
    
    return HttpResponse("success")
    
@csrf_exempt
def test_ajax(request):
    print(request.GET)
    print(request.POST)
    data_dict = {"status": True, 'data': [11, 22, 33, 44]}
    # json_string = json.dumps(data_dict)
    # return HttpResponse(json_string)
    return JsonResponse(data_dict)
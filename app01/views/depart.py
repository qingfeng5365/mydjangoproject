from django.shortcuts import render,redirect,HttpResponse
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse

import os
from openpyxl import load_workbook



from app01 import models
from app01.utils.pagination import Pagination

def depart_list(request):
    # """部门列表"""
    # #queryset = models.Department.objects.all().order_by("id")
    # #queryset = models.Department.objects.all().order_by("-id")

    # queryset = models.Department.objects.all()
    # return render(request,'depart_list.html',{'queryset':queryset})

    conditions_list = {}

    '''
    models.UserInfo.objects.filter(id=12)
    models.UserInfo.objects.filter(id__gt=12)
    models.UserInfo.objects.filter(id__gre=12)
    models.UserInfo.objects.filter(id__lt=12)
    models.UserInfo.objects.filter(id__lte=12)

    models.UserInfo.objects.filter(name__contains="张")
    models.UserInfo.objects.filter(name__startwith="张")
    models.UserInfo.objects.filter(name__endwith="张")
    '''


    search = request.GET.get("search","")
    if search:
        conditions_list["title__contains"] = search
    queryset = models.Department.objects.filter(**conditions_list).order_by("id")

    #分页
    
    page_object = Pagination(request, queryset)
    context = {
        'search':search,
        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()       # 生成页码
        }
    #return render(request,'user_list.html',{'queryset':queryset,'search':search})
    return render(request, 'depart_list.html', context)

def depart_jump(request):

    return JsonResponse(request.GET)

def depart_add(request):
    """新建部门"""
    if request.method == "GET":
        return render(request,'depart_add.html')
    
    title = request.POST.get("title")
    exists = models.Department.objects.filter(title=title).exists()
    if exists:
        error = "部门已存在"
        return render(request,'error.html',{"error":error})
    models.Department.objects.create(title=title)
    return redirect("/depart/list/")

def depart_delete(request):
    """删除部门"""
    nid = request.GET.get("nid")
    exists = models.Department.objects.filter(id=nid).exists()
    if not exists:
        error = "id不存在"
        return render(request,"error.html",{"error":error})
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")

def depart_edit(request,nid):
    """编辑部门"""
    if request.method == "GET":

        exists = models.Department.objects.filter(id=nid).exists()
        if not exists:
            error = "id不存在"
            return render(request,"error.html",{"error":error})

        row_object = models.Department.objects.filter(id=nid).first()
        return render(request,'depart_edit.html',{'row_object':row_object})
    title = request.POST.get("title")
    if title == "":
        error = "名称不能为空"
        return render(request,'error.html',{"error":error})
    exists = models.Department.objects.exclude(id=nid).filter(title=title).exists()
    if exists:
        error = "部门已存在"
        return render(request,'error.html',{"error":error})
    if title==models.Department.objects.filter(id=nid).first().title:
        return redirect("/depart/list/")
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")

@csrf_exempt
def depart_batchAdd(request):
    #print(request.POST)
    #print(request.FILES)
    file_object = request.FILES.get("file")
    #print(file_object.name)

    if file_object.name.split(".")[-1] != "xlsx":
        error = "请上传.xlsx文件"
        data_dict = {"status": False,"error": error}
        return JsonResponse(data_dict)

    # excel文件
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    cell = sheet.cell(1,1)
    #print(cell.value)
    
    if cell.value != "title":
        error = "excel文件的第一个列名必须为title"
        data_dict = {"status": False,"error": error}
        return JsonResponse(data_dict)
        
    for row in sheet.iter_rows(min_row=2):
        title = row[0].value
        exists = models.Department.objects.filter(title=title).exists()
        if exists:
            error = "{}已存在".format(title)
            data_dict = {"status": False,"error": error}
            return JsonResponse(data_dict)
        models.Department.objects.create(title=title)
    data_dict = {"status": True}
    return JsonResponse(data_dict)

@csrf_exempt
def depart_batchDelete(request):
    #print(request.POST)
    nid_list = request.POST.getlist("nid_list")
    #print(nid_list)
    if not nid_list:
        error = "请选中需要删除的id"
        data_dict = {"status": False,"error":error}
        return JsonResponse(data_dict)
    for nid in nid_list:
        exists = models.Department.objects.filter(id=nid).exists()
        if not exists:
            error = "id不存在"
            data_dict = {"status": False,"error":error}
            return JsonResponse(data_dict)
        models.Department.objects.filter(id=nid).delete()
    data_dict = {"status": True}
    return JsonResponse(data_dict)
    

         

